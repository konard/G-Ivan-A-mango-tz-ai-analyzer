"""Generate placeholder knowledge_base source files.

Creates:
    - knowledge_base/sources/functional_list_sample.xlsx (10-15 rows)
    - knowledge_base/sources/api_guide_excerpt.pdf (1-2 pages)
    - knowledge_base/sources/user_guide_call_recording.pdf (1 page)

Run from repo root:

    python3 scripts/generate_kb_sources.py
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer


SOURCES_DIR = Path(__file__).resolve().parents[1] / "knowledge_base" / "sources"

FUNCTIONAL_HEADERS = ["ID", "Название", "Описание", "Статус"]
FUNCTIONAL_ROWS = [
    ("F001", "Запись разговоров", "Сохранение всех входящих и исходящих звонков в облачное хранилище.", "Active"),
    ("F002", "Голосовое меню (IVR)", "Многоуровневое голосовое меню с переадресацией по добавочным номерам.", "Active"),
    ("F003", "Интеграция с Битрикс24", "Двусторонняя синхронизация контактов и сделок через коннектор v2.1.", "Active"),
    ("F004", "Интеграция с amoCRM", "Прокидывание звонков и карточек клиентов в amoCRM.", "Active"),
    ("F005", "Аналитика звонков", "Отчёты по длительности, пропущенным звонкам и нагрузке операторов.", "Active"),
    ("F006", "Анализ эмоций (постобработка)", "Определение тональности разговора после завершения звонка.", "Beta"),
    ("F007", "Транскрибация речи", "Перевод аудио в текст с распознаванием говорящего.", "Active"),
    ("F008", "Public API", "REST API для управления звонками и получением событий.", "Active"),
    ("F009", "SIP-транк", "Подключение собственной АТС по SIP-протоколу.", "Active"),
    ("F010", "Виджет обратного звонка", "JS-виджет для сайта, заказывающий звонок клиенту.", "Active"),
    ("F011", "Контроль качества", "Чек-листы и тегирование звонков супервизором.", "Active"),
    ("F012", "Мобильное приложение", "Клиент для iOS/Android с push-уведомлениями.", "Active"),
    ("F013", "Очередь звонков", "Распределение входящих звонков по группам операторов.", "Active"),
    ("F014", "Запись экрана оператора", "Видеозапись рабочего стола во время разговора.", "Beta"),
    ("F015", "WebRTC-софтфон", "Браузерный софтфон без установки приложений.", "Active"),
]

API_GUIDE_PARAGRAPHS = [
    ("Platform API Documentation — Excerpt", "Heading1"),
    (
        "Документ описывает публичные методы API целевой платформы. "
        "Все методы используют HTTPS и аутентификацию по api_key + sign (HMAC-SHA256).",
        "BodyText",
    ),
    ("POST /call/start", "Heading2"),
    (
        "Инициирует исходящий звонок с указанного добавочного номера на внешний номер. "
        "Параметры запроса: from (extension), to (E.164), command_id. "
        "Метод описан в Public API Guide v2.0.",
        "BodyText",
    ),
    ("POST /stats/request", "Heading2"),
    (
        "Асинхронный запрос статистики по звонкам за период. "
        "Возвращает request_id, по которому позже забираются результаты.",
        "BodyText",
    ),
    ("Webhook events", "Heading2"),
    (
        "Сервис может присылать события call_event, summary, recording по адресу подписчика. "
        "Подпись проверяется через тот же ключ HMAC-SHA256.",
        "BodyText",
    ),
]

CALL_RECORDING_PARAGRAPHS = [
    ("Запись разговоров — Руководство пользователя", "Heading1"),
    (
        "Запись разговоров доступна на всех тарифах уровня Профи и выше. "
        "Записи хранятся в облаке целевой платформы до 3 лет с возможностью продления.",
        "BodyText",
    ),
    ("Как включить запись", "Heading2"),
    (
        "В личном кабинете перейдите в раздел «Сотрудники» → выберите добавочный номер → "
        "включите переключатель «Запись разговоров». Запись стартует автоматически при "
        "следующем звонке.",
        "BodyText",
    ),
    ("Где скачать запись", "Heading2"),
    (
        "Записи доступны в разделе «Статистика и записи». Также их можно получать через "
        "API метод GET /recording/{id} или webhook recording.",
        "BodyText",
    ),
]


def build_functional_list(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Functional list"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")

    ws.append(FUNCTIONAL_HEADERS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in FUNCTIONAL_ROWS:
        ws.append(row)

    widths = [10, 32, 70, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = wrap

    wb.save(path)


def build_pdf(path: Path, blocks: list[tuple[str, str]]) -> None:
    doc = SimpleDocTemplate(str(path), pagesize=A4, title=path.stem)
    styles = getSampleStyleSheet()
    story = []
    for text, style in blocks:
        story.append(Paragraph(text, styles[style]))
        story.append(Spacer(1, 8))
    doc.build(story)


def main() -> None:
    SOURCES_DIR.mkdir(parents=True, exist_ok=True)

    xlsx_path = SOURCES_DIR / "functional_list_sample.xlsx"
    build_functional_list(xlsx_path)
    print(f"Saved: {xlsx_path}")

    api_pdf = SOURCES_DIR / "api_guide_excerpt.pdf"
    build_pdf(api_pdf, API_GUIDE_PARAGRAPHS)
    print(f"Saved: {api_pdf}")

    rec_pdf = SOURCES_DIR / "user_guide_call_recording.pdf"
    build_pdf(rec_pdf, CALL_RECORDING_PARAGRAPHS)
    print(f"Saved: {rec_pdf}")


if __name__ == "__main__":
    main()
